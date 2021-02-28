import requests
import Secrets
import sqlite3
from openpyxl import load_workbook
from typing import Tuple


def open_db(filename: str) -> Tuple[sqlite3.Connection, sqlite3.Cursor]:
    db_connection = sqlite3.connect(filename)  # connect to existing DB or create new one
    cursor = db_connection.cursor()  # get ready to read/write data
    return db_connection, cursor


def close_db(connection: sqlite3.Connection):
    connection.commit()  # make sure any changes get saved
    connection.close()


def setup_db(cursor: sqlite3.Cursor):
    cursor.execute('''CREATE TABLE IF NOT EXISTS university_data(
    school_id INTEGER PRIMARY KEY,
    university_name TEXT NOT NULL,
    university_state TEXT NOT NULL,
    university_city TEXT,
    student_2018 INT,
    three_year_earnings_over_poverty INT,
    loan_repayment INT);''')


def setup_excel_db(cursor: sqlite3.Cursor):
    cursor.execute('''CREATE TABLE IF NOT EXISTS wage_and_job(
    state TEXT
    occupation_major_title TEXT
    total_employment_in_field INT
    25th_percentile_salary INT
    occupation_code INT
    );''')


def insert_data_excel(excel_data, cursor):
    return


def insert_data(all_data, cursor):
    for univ_data in all_data:
        cursor.execute("""
        INSERT INTO university_data(school_id, university_name, student_2018, university_state, 
        three_year_earnings_over_poverty, loan_repayment)
         VALUES (?,?,?,?,?,?);
        """, (univ_data['id'], univ_data['school.name'], univ_data['2018.student.size'], univ_data['school.state'],
              univ_data['2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line'],
              univ_data['2016.repayment.3_yr_repayment.overall']))


def get_data_excel(row_count):
    workbook = load_workbook(filename="state_M2019_dl.xlsx")
    workbook.sheetnames
    sheet = workbook.active
    excel_state_list = []
    excel_occupation_major_title = []
    excel_total_employment = []
    excel_25th_percentile = []
    excel_occupation_code = []

    for value in sheet.iter_rows(min_row=2, max_row=row_count, min_col=2, max_col=2, values_only=True):
        excel_state_list.append(value)
    for value in sheet.iter_rows(min_row=2, max_row=row_count, min_col=9, max_col=9, values_only=True):
        excel_occupation_major_title.append(value)
    for value in sheet.iter_rows(min_row=2, max_row=row_count, min_col=11, max_col=11, values_only=True):
        excel_total_employment.append(value)
    for value in sheet.iter_rows(min_row=2, max_row=row_count, min_col=20, max_col=20, values_only=True):
        excel_25th_percentile.append(value)
    for value in sheet.iter_rows(min_row=2, max_row=row_count, min_col=8, max_col=8, values_only=True):
        excel_occupation_code.append(value)

    return excel_state_list, excel_occupation_code, excel_total_employment, excel_25th_percentile, excel_occupation_code


def get_data():
    all_data = []
    response = requests.get(f'https://api.data.gov/ed/collegescorecard/v1/schools.json?'
                            f'school.degrees_awarded.predominant=2,3&fields=id,school.state,school.name,'
                            f'2018.student.size,2016.repayment.3_yr_repayment.overall,'
                            f'2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line'
                            f'&api_key={Secrets.api_key}')
    first_page = response.json()
    if response.status_code != 200:
        print(F"Error getting data from API: {response.raw}")
        return []
    total_results = first_page['metadata']['total']
    page = 0
    per_page = first_page['metadata']['per_page']
    all_data.extend(first_page['results'])
    while (page + 1) * per_page < total_results:
        page += 1
        response = requests.get(
            f'https://api.data.gov/ed/collegescorecard/v1/schools.json?school.degrees_awarded.predominant=2,3'
            f'&fields=id,school.state,school.name,2018.student.size,2016.repayment.3_yr_repayment.overall,'
            f'2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line'
            f'&api_key={Secrets.api_key}&page={page}')
        if response.status_code != 200:
            continue
        current_page = response.json()
        all_data.extend(current_page['results'])

    return all_data


def main():
    row_count = 36383  # total number of rows in excel
    # all_data = get_data()
    # conn, cursor = open_db("Schools_Database.sqlite")
    # setup_db(cursor)
    # insert_data(all_data, cursor)
    # close_db(conn)
    excel_state = []
    excel_occupation_title = []
    excel_total_employment = []
    excel_25th_percent = []
    excel_occupation_code = []
    excel_state, excel_occupation_title, excel_total_employment, excel_25th_percent, excel_occupation_code =\
        get_data_excel(row_count)
    print(excel_total_employment)


if __name__ == '__main__':
    main()
